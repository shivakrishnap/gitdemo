import openpyxl


class Hompagedata:
    test_homepage_data = [{"firstname": "shiva", "email": "shivakrishna@gmail.com", "gender": "Male"},{"firstname": "sathya", "email": "sathyakrishna@gmail.com", "gender": "Female"}]

    @staticmethod
    def getstdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\shiva\\OneDrive\\Documents\\pyexceldemo.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        print(cell.value)

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i,column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
