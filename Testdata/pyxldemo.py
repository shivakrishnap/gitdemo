import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\shiva\\OneDrive\\Documents\\pyexceldemo.xlsx")
sheet = wb.active
sheet.cell(row=1, column=5).value = "hello"
print(sheet.cell(row=1, column=5).value)
wb.save('pyexceldemo.xlsx')
